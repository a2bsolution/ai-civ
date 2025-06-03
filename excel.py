import pandas as pd
import numpy as np
import re
from functions import container_separate, clean_amount

import argparse
from sentence_transformers import SentenceTransformer, util
from typing import Dict, List, Optional, Tuple

import pytesseract
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io


# Target fields and prompts for semantic scoring
TARGET_FIELDS = {
    "invoice_number": "ci no",
    "issued_date": "issued date",
    "container_number": "container",
    "discount": "discount rebate",
    "total": "total invoice amount"
}

FIELD_ALIASES = {
    "invoice_number": ["invoice number", "ci no", "invoice no"],
    "issued_date": ["issued date", "date"],
    "container_number": ["container"],
    "discount": ["discount", "rebate"],
    "total": ["invoice amount", "total", "invoice total"],
}


TABLE_COLUMNS = [
    "product_code",
    "goods_description",
    "invoice_quantity",
    "unit_price",
    "price"
]

COLUMN_ALIASES = {
    "product_code": ["product code","item no", "item code", "#", "item no.", "nsitem no"],
    "goods_description": ["description", "item description"],
    "invoice_quantity": ["qty", "quantity"],
    "unit_price": ["unit price", "price/unit", "rate", "price each"],
    "price": ["amount", "line amt", "line amount", "total", "total value"]
}

CANONICAL_LINE_ITEM_HEADERS = [
    {"product code", "item description", "price", "quantity", "total value"},
    {"item no", "description", "qty", "unit price", "total value", "total gross weight"},
    {"qty", "nsitem no", "description", "line amt"}
]

model = SentenceTransformer("all-MiniLM-L6-v2")

def incoterm_filter(incoterm):
    incoterm = re.sub('[^A-Za-z ]+', '', incoterm).upper()
    incoterm_list = ['FOB', 'CFR', 'CIF', 'CIP', 'CPT', 'DAP', 'DAT', 'DDP', 'DPU', 'EXW', 'FAS', 'FC1', 'FC2', 'FCA']

    for word in incoterm_list:
        if word in incoterm:
            return word
            break
    return None

def excel_contains_image(file_bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes))
    first_sheet = wb.worksheets[0]
    for image in first_sheet._images:
        if is_image_near_top(image):
            return bool(getattr(first_sheet, "_images", []))
            break
        else:
            print("[DEBUG] Ignoring image — not near top")

def is_image_near_top(image, max_row=6):
    """
    Determines if the image appears near the top of the Excel sheet.
    This relies on the anchor cell reference like 'A1', 'B2', etc.
    """
    try:
        anchor = image.anchor._from if hasattr(image.anchor, "_from") else image.anchor
        row = getattr(anchor, 'row', 0)  # 0-based
        return row < max_row
    except Exception as e:
        print(f"[WARN] Could not determine image position: {e}")
        return False


def extract_supplier_from_excel_image(file_path):
    wb = load_workbook(file_path)
    first_sheet = wb.worksheets[0]
    supplier_text = None

    for img in getattr(first_sheet, "_images", []):
        # Read the image bytes and convert to PIL
        img_bytes = img._data()
        img_pil = PILImage.open(io.BytesIO(img_bytes))

        # OCR the image
        text = pytesseract.image_to_string(img_pil).upper()
        #print(f"[OCR IMAGE TEXT]\n{text}")

        # Look for likely supplier name
        for line in text.splitlines():
            if any(kw in line for kw in ["LIMITED", "LTD", "CORP", "INC", "PTE"]):
                supplier_text = line.strip().title()
                print(f"[MATCH] supplier ← OCR image → '{supplier_text}'")
                return supplier_text

    return None


# Helper to clean and extract value after colon, dash, etc.
def extract_value(cell_text):
    parts = re.split(r":", cell_text, maxsplit=1)
    if len(parts) > 1:
        return parts[1].strip()
    return cell_text.strip()


# Clean and convert various date formats
def clean_date(val):
    try:
        for fmt in ("%d-%b-%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        # fallback
        date = pd.to_datetime(val, errors="coerce", dayfirst=True)
        return date.strftime("%Y-%m-%d") if pd.notnull(date) else val
    except:
        return val

incoterm_list = ['FOB', 'CFR', 'CIF', 'CIP', 'CPT', 'DAP', 'DAT', 'DDP', 'DPU', 'EXW', 'FAS', 'FC1', 'FC2', 'FCA']

def detect_incoterm_from_df(df: pd.DataFrame, incoterm_list: List[str]) -> Optional[str]:
    allowed_terms = ["PRICE TERM", "DELIVERY TERM", "INCOTERM", "TERMS OF DELIVERY"]
    df_str = df.astype(str).map(lambda x: x.strip().upper())

    for row in df_str.itertuples(index=False):
        row = list(row)
        for i in range(len(row)):
            key = row[i]
            if not isinstance(key, str):
                continue

            if any(term in key.upper() for term in allowed_terms):
            # Look ahead for up to 3 cells
                for j in range(i + 1, min(i + 4, len(row))):
                    val = row[j]
                    if isinstance(val, str) and val.strip():
                        clean_val = val.strip().upper()
                        for incoterm in incoterm_list:
                            if incoterm in clean_val:
                                print(f"[MATCH] incoterm ← '{key}' → '{val}'")
                                return incoterm
                        break


    for row in df_str.itertuples(index=False):
        for cell in row:
            # Skip cells from known misleading labels
            if "ISSUED DATE" in cell or "DATE" in cell:
                continue

            for incoterm in incoterm_list:
                if cell.startswith(incoterm):
                    print(f"[MATCH:FUZZY] incoterm ← full cell → '{cell}'")
                    return incoterm

    return None


def normalize_header(text):
    text = str(text)
    text = re.sub(r"[^\w\s]", "", text)  # Remove punctuation
    text = re.sub(r"\s+", " ", text)      # Collapse multiple spaces into single space
    return text.strip().lower()

def detect_line_item_header_row(df, line_item_headers=CANONICAL_LINE_ITEM_HEADERS):
    for i, row in df.iterrows():
        normalized_row = [normalize_header(cell) for cell in row]

        for canonical_set in line_item_headers:
            # Normalize canonical fields too
            normalized_canon = {normalize_header(field) for field in canonical_set}
            match_count = sum(
                any(
                    cell == field or
                    field in cell.split() or
                    cell in field.split()
                    for cell in normalized_row
                )
                for field in normalized_canon
            )

            if match_count >= len(normalized_canon) // 2:
                print(f"[DEBUG] Header row index: {i} with {match_count} column matches")
                return i
    print("[DEBUG] No matching header row found")
    return None

def remap_headers(columns):
    mapped = {}

    columns_lower = [normalize_header(col) for col in columns]

    # Special logic for product_code
    if any("product code" in col for col in columns_lower):
        for col in columns:
            if "product code" in normalize_header(col):
                mapped["product_code"] = col
                break
    elif any("item no" in col for col in columns_lower):
        for col in columns:
            if "item no" in normalize_header(col):
                mapped["product_code"] = col
                break

    # Map the rest normally
    for target, aliases in COLUMN_ALIASES.items():
        if target == "product_code":
            continue  # Already handled separately above
        for col in columns:
            if any(alias in normalize_header(col) for alias in aliases):
                mapped[target] = col
                break

    return mapped


def detect_currency_from_headers(headers):
    currency_map = {"usd": "USD", "aud": "AUD", "eur": "EUR", "jpy": "JPY"}
    for col in  headers:
        col_lower = str(col).lower()
        for symbol, iso in currency_map.items():
            if symbol in col_lower:
                currency = iso
                print(f"[INFO] Detected currency: {currency} from column: '{col}'")
                return currency

def detect_incoterm_from_headers(headers: List[str], incoterm_list=incoterm_list) -> Optional[str]:
    for header in headers:
        header_norm = header.strip().upper()
        for term in incoterm_list:
            if term in header_norm:
                print(f"[INFO] Incoterm detected from header: '{header}' → '{term}'")
                return term
    return None

def extract_supplier(df: pd.DataFrame) -> str:
    for row in df.itertuples(index=False):
        for val in row:
            if pd.notna(val) and str(val).strip():
                print(f"[MATCH] supplier ← fallback cell → '{val}'")
                return str(val).strip()
    return None

def try_parse(val):
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except:
        return None

def extract_table(df: pd.DataFrame, header_idx: int) -> Tuple[Dict[str, List], Optional[float], Optional[float]]:
    df_items = df.iloc[header_idx + 1:].copy()
    df_items.reset_index(drop=True, inplace=True)

    normalized_cols = [normalize_header(col) for col in df.iloc[header_idx]]
    df_items.columns = normalized_cols
    df_items = df_items.dropna(how="all")

    mapped_cols = remap_headers(normalized_cols)
    print(f"[DEBUG] Mapped columns: {mapped_cols}")

    for std_name, original_col in mapped_cols.items():
        df_items.rename(columns={original_col: std_name}, inplace=True)

    table = {k: [] for k in ["product_code", "goods_description", "invoice_quantity", "unit_price", "price"]}
    termination_keywords = {"invoice amount", "total amount", "line total", "total"}

    total = None
    total_weight = None
    stop_idx = None
    # Try to detect currency from headers (e.g., "Line Amt. (USD)", "Total Value EUR")
    currency = None
    currency_candidates = [col for col in df.iloc[header_idx] if isinstance(col, str)]

    # Pattern for ISO currency codes in parentheses
    code_pattern = re.compile(r"\((USD|AUD|EUR|JPY)\)", re.IGNORECASE)
    # Pattern for symbols
    symbol_map = {"$": "USD", "€": "EUR", "¥": "JPY", "A$": "AUD"}

    for col in currency_candidates:
        match_code = code_pattern.search(col)
        if match_code:
            currency = match_code.group(1).upper()
            print(f"[INFO] Found currency from header: {currency}")
            break

    # Fallback: symbol detection from header
    if not currency:
        for col in currency_candidates:
            for symbol, iso in symbol_map.items():
                if symbol in col:
                    currency = iso
                    print(f"[INFO] Inferred currency from symbol '{symbol}' in header: {currency}")
                    break
            if currency:
                break


    for idx, row in df_items.iterrows():
        row_cells = [normalize_header(str(cell)) for cell in row]

        if any(keyword in cell for cell in row_cells for keyword in termination_keywords):
            print("[DEBUG] Stopping table parse at summary row:", row_cells)

            # Extract total if price column is available
            if "price" in row.index:
                total_val = try_parse(row["price"])
                if total_val:
                    total = round(total_val, 2)
                    print(f"[INFO] Found total = {total}")

            # Extract total_weight if 'total_weight' column is already renamed
            weight_col = next(
                (col for col in df_items.columns if "gross weight" in normalize_header(col) and "total" in normalize_header(col)),
                None
            )
            if weight_col and weight_col in row.index:
                weight_val = try_parse(row[weight_col])
                if weight_val:
                    total_weight = round(weight_val, 2)
                    print(f"[INFO] Found total_weight = {total_weight}")

            stop_idx = idx
            break

        # Case 2: Fallback — if "inland charge" found
        if any("inland charge" in cell for cell in row_cells):
            print("[DEBUG] Inland charge detected. Looking for preceding summary row...")

            inland_idx = idx
            for rev_offset in range(1, 6):
                if inland_idx - rev_offset < 0:
                    break
                row_above = df_items.iloc[inland_idx - rev_offset]
                if "price" in row_above.index:
                    total_val = try_parse(row_above["price"])
                    if total_val:
                        total = round(total_val, 2)
                        print(f"[INFO] Extracted total from summary row = {total}")
                weight_col = next((col for col in df_items.columns if all(kw in normalize_header(col) for kw in ["gross", "weight", "total"])),None)
                if weight_col:
                    weight_val = try_parse(row_above.get(weight_col))
                    if weight_val:
                        total_weight = round(weight_val, 2)
                        print(f"[INFO] Extracted total_weight from summary row = {total_weight}")
                if total_val or weight_val:
                    stop_idx = df_items.index[inland_idx - rev_offset]
                    break

    # Table parsing loop
    if stop_idx is not None:
        limit = df_items.index.get_loc(stop_idx)
        iterator = df_items.iloc[:limit].iterrows()
    else:
        iterator = df_items.iterrows()

    for _, row in iterator:
        prod = str(row.get("product_code", "")).strip()
        desc = str(row.get("goods_description", "")).strip()
        qty = try_parse(row.get("invoice_quantity", ""))
        price = try_parse(row.get("unit_price", ""))
        amount = try_parse(row.get("price", ""))

        if prod.strip("_").strip().lower() == "" and all(x in [None, "", 0.0] for x in [desc, qty, price, amount]):
            continue

        table["product_code"].append(prod)
        table["goods_description"].append(desc)
        table["invoice_quantity"].append(qty)
        table["unit_price"].append(price)
        table["price"].append(amount)

    return table, total, total_weight, currency


# Enhanced row-based field extractor
def extract_fields_by_row(df, threshold):
    results = {key: None for key in TARGET_FIELDS}
    results["incoterm"] = None  # NEW
    used_cells = set()
    df = df.fillna(method='ffill', axis=1)  # Handle merged cells

    for row_idx, row in df.iterrows():
        cells = [str(cell).strip() for cell in row if str(cell).strip()]
        if len(cells) < 2:
            continue

        for i in range(len(cells) - 1):
            key_cell = cells[i].lower()
            val_cell = cells[i + 1].strip()

            # Priority 1: Exact FIELD_ALIASES match
            for key, aliases in FIELD_ALIASES.items():
                if results[key]:
                    continue
                if any(normalize_header(alias) == normalize_header(key_cell) for alias in aliases):
                    print(f"[MATCH:ALIAS] {key} ← '{key_cell}' → '{val_cell}'")
                    if val_cell in used_cells:
                        continue
                    used_cells.add(val_cell)
                    results[key] = post_process_field(key, val_cell)
                    break

        # Fallback: Fuzzy matching if still missing
        for i in range(len(cells) - 1):
            key_cell = cells[i].lower()
            val_cell = cells[i + 1].strip()

            for key, label in TARGET_FIELDS.items():
                if results.get(key):
                    continue
                prompt_emb = model.encode(label, convert_to_tensor=True)
                key_emb = model.encode(key_cell, convert_to_tensor=True)
                similarity = float(util.cos_sim(prompt_emb, key_emb))

                if  similarity >= threshold and val_cell != key_cell:
                    if val_cell in used_cells:
                        continue
                    used_cells.add(val_cell)
                    print(f"[MATCH:FUZZY] {key} ← '{key_cell}' → '{val_cell}' (score: {similarity:.3f})")
                    results[key] = post_process_field(key, val_cell)

        # Additional scan for discount
        for i in range(len(cells) - 1):
            key_cell = cells[i].lower()
            val_cell = cells[i + 1].strip()

            if key == "discount":
                numeric = re.search(r"-?\$?\d{1,3}(,\d{3})*(\.\d{2})", val_cell)
                if numeric:
                    results[key] = clean_amount(numeric.group())

        # Incoterm match
        if not results["incoterm"]:
            results["incoterm"] = detect_incoterm_from_df(df, incoterm_list)

    return results

def post_process_field(key, val_cell):
    if key == "issued_date":
        return clean_date(val_cell)
    elif key == "total" and re.search(r"\$?\d{1,3}(,\d{3})*(\.\d{2})?$", val_cell):
        return clean_amount(val_cell)
    elif key == "container_number" and re.search(r"[A-Z]{4}\d{7}$", val_cell):
        return val_cell
    elif key == "discount":
        m = re.search(r"-?\$?\d{1,3}(,\d{3})*(\.\d{2})", val_cell)
        return clean_amount(m.group()) if m else None
    return extract_value(val_cell)


def match_column_headers(headers):
    headers_lower = [h.lower().strip() for h in headers]
    embeddings = model.encode(headers_lower, convert_to_tensor=True)
    target_embeddings = model.encode(TABLE_COLUMNS, convert_to_tensor=True)
    matched_columns = {}

    for i, target in enumerate(TABLE_COLUMNS):
        sims = util.cos_sim(target_embeddings[i], embeddings)[0]
        best_score = float(sims.max())
        best_index = int(sims.argmax())
        best_match = headers[best_index]

        # If high confidence, use it
        if best_score >= 0.6:
            matched_columns[target] = best_match
            print(f"[MATCH:AI] '{target}' → '{best_match}' (score: {best_score:.3f})")
        else:
            # Fallback to alias matching
            aliases = COLUMN_ALIASES.get(target, [])
            for j, header in enumerate(headers_lower):
                if any(alias in header for alias in aliases):
                    matched_columns[target] = headers[j]
                    print(f"[MATCH:ALIAS] '{target}' → '{headers[j]}' (alias match)")
                    break
            else:
                matched_columns[target] = None  # No match

    return matched_columns


def is_number(val):
    """Check if a value is a number (after removing commas)."""
    try:
        float(str(val).replace(",", ""))
        return True
    except:
        return False

def extract_total_weight_from_last_row(df: pd.DataFrame) -> Optional[float]:
    LINE_ITEM_HEADERS = [{"product code", "item description", "pcs", "cbm", "total weight"}]
    header_idx = detect_line_item_header_row(df, LINE_ITEM_HEADERS)
    df.columns = df.iloc[header_idx]
    df = df.iloc[header_idx + 1:].reset_index(drop=True)

    normalized_col_names = [normalize_header(col) for col in df.columns]
    
    # First pass: prioritize columns with 'gross' or 'gw'
    priority_keywords = ["total weightgw", "gross", "gw"]
    fallback_keywords = ["weight"]  # general weight check only if no match in priority

    for keyword_group in [priority_keywords, fallback_keywords]:
        for idx in reversed(df.index):
            row = df.loc[idx]
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip()
                col_header = normalized_col_names[col_idx]
                if any(kw in col_header for kw in keyword_group) and is_number(cell_str):
                    weight = try_parse(cell)
                    if weight:
                        print(f"[INFO] Fallback: Found total_weight with keyword group {keyword_group} = {weight}")
                        return round(weight, 2)
    return None


def extract_total_weight_from_sheet(file_bytes, sheet_name):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, dtype=str).fillna("")
    weight = extract_total_weight_from_last_row(df)
    if weight:
        return weight
    return None

def extract_xlsx(file_bytes, threshold=0.6):
    bytes_io = io.BytesIO(file_bytes)
    wb = load_workbook(bytes_io, read_only=True)
    sheet_names = [name.strip().upper() for name in wb.sheetnames]

    main_sheet = next((name for name in wb.sheetnames if name.strip().upper() == "IV"), wb.sheetnames[0])

    df = pd.read_excel(bytes_io, sheet_name=main_sheet, header=None, dtype=str).fillna("")
    prediction = extract_fields_by_row(df, threshold)

    # Try to find the table header row using AI match score density
    df_str = df.astype(str).apply(lambda x: x.str.strip().str.lower(), axis=1)
    header_row_idx = None
    max_matches = 0

    for i, row in df_str.iterrows():
        headers = row.tolist()
        embeddings = model.encode(headers, convert_to_tensor=True)
        targets = model.encode(TABLE_COLUMNS, convert_to_tensor=True)

        matches = 0
        for t in targets:
            sims = util.cos_sim(t, embeddings)[0]
            if float(sims.max()) > 0.55:
                matches += 1

        if matches > max_matches:
            max_matches = matches
            header_row_idx = i

    prediction["supplier"] = None  # Add field to the result structure
    prediction["page"] = [1]

    # Only extract supplier from image if image is detected
    if excel_contains_image(file_bytes):
        print("[INFO] Image detected in Excel — running OCR for supplier")
        supplier = extract_supplier_from_excel_image(io.BytesIO(file_bytes))
        if supplier:
            prediction["supplier"] = supplier
    else:
        prediction["supplier"] = extract_supplier(df)

    header_idx = detect_line_item_header_row(df)
    if header_idx is not None:
        currency = detect_currency_from_headers(df.iloc[header_idx])
        prediction["currency"] = currency

        prediction["incoterm"] = prediction["incoterm"] or detect_incoterm_from_headers(df.iloc[header_idx])
        table, total, total_weight, currency = extract_table(df, header_idx)
        prediction["table"] = table
        if total_weight is not None:
            prediction["total_weight"] = total_weight
        if currency is not None:
            prediction["currency"] = currency
        #table_df['unit_quantity'] = ["Piece"] * len(table_df['invoice_quantity'])

        #prioritize table found from extract_table, then one from fields, then fallback
        if total is not None:
            prediction["total"] = total
        else:
            if prediction.get("total"):
                pass
            else:
                try:
                    prediction["total"] = round(sum(p for p in prediction["table"]["price"] if p), 2)
                    print(f"[INFO] Fallback total = {prediction['total']}")
                except Exception as e:
                    print(f"[WARN] Failed to compute fallback total: {e}")

    else:
        prediction["table"] = {key: [] for key in ["product_code", "goods_description", "invoice_quantity", "unit_price", "price"]}


    if "PL" in sheet_names:
        weight = extract_total_weight_from_sheet(file_bytes, "PL")
        if weight:
            prediction["total_weight"] = weight  # overwrite or add

    return prediction

def extract_xls(df):
    prediction = {'page': [1], 'invoice_number': None, 'incoterm': None, 'container': None}

    prediction['supplier'] = next((val for row in df.itertuples(index=False) for val in row if pd.notna(val)), None)

    # Iterate through the DataFrame to find the necessary values
    for row in df.iterrows():
        row_idx, row_data = row
        for col_idx, value in enumerate(row_data):
            if isinstance(value, str):
                value_clean = value.strip().lower()

                if "invoice no" in value_clean:
                    prediction['invoice_number'] = df.iloc[row_idx, col_idx + 1]  # Get the value next to "Invoice No."
                
                elif "price term" in value_clean:
                    col_add = 1
                    while col_add < 5:
                        value = df.iloc[row_idx, col_idx + col_add] 
                        if pd.notna(value) and value != '':
                            prediction['incoterm'] = value
                            break
                        col_add += 1

                    if pd.isna(prediction['incoterm']):
                        prediction['incoterm'] = incoterm_filter(str(df.iloc[row_idx, col_idx + 2]).split()[0])
                    else:
                        prediction['incoterm'] = incoterm_filter(str(prediction['incoterm']).split()[0])

                elif "container" in value_clean:
                    if col_idx + 1 < df.shape[1] and pd.notna(df.iloc[row_idx, col_idx + 1]) and str(df.iloc[row_idx, col_idx + 1]).strip():
                        filtered_container = container_separate(df.iloc[row_idx, col_idx + 1])
                        if filtered_container:
                            prediction['container'] = filtered_container[0]
                        else:
                            prediction['container'] = df.iloc[row_idx, col_idx + 1]

            # If both values are found, exit the loop early
        if prediction['invoice_number'] is not None and prediction['incoterm'] is not None and prediction['container'] is not None:
            break

    item_no_positions = df.apply(lambda row: row.astype(str).str.contains("ITEM NO", case=False, na=False), axis=1)
    first_item_no_position = item_no_positions.stack()
    header_row, header_col = first_item_no_position[first_item_no_position].index[0]
    
    if not first_item_no_position.empty:
        table_headers = df.iloc[header_row]
        table_df = df.iloc[header_row + 1:]  # Extract the data from the rows below the headers

        normalized_headers = table_headers.astype(str).str.strip().str.lower()
        matching_columns = normalized_headers[normalized_headers.str.contains("total gross weight", na=False)]
        total_weight_col_idx = matching_columns.index[0] if not matching_columns.empty else -1

        prediction['currency'] = df.iloc[header_row, 4]

        table_df = table_df.iloc[:, header_col:header_col+5]
        table_df = table_df.dropna(how='all')  # Drop rows that are entirely empty

        for index, row in table_df.iterrows():
            if isinstance(row.iloc[0], str) and "inland charge" in row.iloc[0].lower():
                table_df = table_df.loc[:index - 1]  # Only keep rows before this point
                last_row_value = re.sub('[^0-9.]+', '', str(table_df.iloc[-1, 4]))  # Get the value in column E (index 4) of the last row
                prediction['total'] = round(float(last_row_value), 2)

                if total_weight_col_idx != -1:
                        total_weight_value = df.iloc[index-2, total_weight_col_idx]
                        prediction['total_weight'] = round(float(re.sub(r'[^0-9.]+', '', str(total_weight_value))), 2)

                break

        if not table_df.empty:
            table_df = table_df.dropna(subset=[table_df.columns[0], table_df.columns[1]])  # Remove NaN values
            table_df = table_df[table_df.iloc[:, 0].str.strip().astype(bool) & table_df.iloc[:, 1].str.strip().astype(bool)] # Remove rows where first or second column is empty or only spaces

        # Set the headers as column names of the table DataFrame
        table_headers = ['product_code', 'goods_description', 'invoice_quantity', 'unit_price', 'price']
        table_df.columns = table_headers
        table_df['unit_quantity'] = ["Piece"] * len(table_df['invoice_quantity'])
        prediction['table'] = table_df.to_dict(orient='list')

    else:
        raise ValueError("The header 'ITEM NO.:' could not be found in the file.")

    return prediction